from dataclasses import dataclass
from aiogram.utils.markdown import hcode
import openpyxl


class Datasheet:

    @dataclass
    class SheetData:
        """stores workbook path and sheet name"""
        wb_path: str
        sheet_name: str

    @dataclass
    class SheetNavigation:
        """stores parameters for specific columns in a datasheet"""
        first_row: int = 2
        name_column: int = 0
        ids_column: int = 1
        relevance_column: int = 2
        keys_column: int = 3
        items_column: int = 4

    def __init__(self, workbook_path, this_sheet_name) -> None:
        self._datasheet_dict = {}
        self._SheetData = self.SheetData(wb_path=workbook_path, sheet_name=this_sheet_name)
        self.SheetNav = self.SheetNavigation
        self.wb = openpyxl.load_workbook(self._SheetData.wb_path)

    def get_ids(self) -> list:
        """Reads IDs from a specific column in the datasheet"""
        sheet = self.wb[self._SheetData.sheet_name]
        sheet_category_ids = []
        for row in range(self.SheetNav.first_row, sheet.max_row):
            if sheet[row][self.SheetNav.ids_column].value:
                sheet_category_ids.append(sheet[row][self.SheetNav.ids_column].value)
        return sheet_category_ids

    def fill_datasheet_dict(self) -> None:
        """fill datasheet from the data in a specific sheet"""
        sheet = self.wb[self._SheetData.sheet_name]
        sheet_keys = []
        sheet_items = []
        for row in range(self.SheetNav.first_row, sheet.max_row):
            row_data = sheet[row]

            if all((
                    row_data[self.SheetNav.keys_column].value,  # Проверка на наличие элемента...
                    row_data[self.SheetNav.items_column].value,
                    row_data[self.SheetNav.relevance_column].value
            )):
                key = row_data[self.SheetNav.keys_column].value.lower().replace("\"", "")
                sheet_keys.append(key)
                sheet_items.append(row_data[self.SheetNav.items_column].value)
                # TODO: Убрать вхардкоженные строки внутри сущности
                item_with_relevance = [hcode(sheet_items[sheet_keys.index(key)]), f"актуальность: {row_data[self.SheetNav.relevance_column].value}"]
                self._datasheet_dict[key] = item_with_relevance

    def access_datasheet_dict(self) -> dict:
        """gives access to datasheet dictionary -> for use in Database"""
        return self._datasheet_dict
